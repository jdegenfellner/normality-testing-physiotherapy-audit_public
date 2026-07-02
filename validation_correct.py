#!/usr/bin/env python3
"""Read the human-coded validation sheet and compute:
  (1) agreement + Cohen's kappa between the human gold standard and the
      rule-based classifier, weighted back to the corpus (the sample is
      stratified, so the raw agreement on the 100 is not corpus-representative);
  (2) a design-based, misclassification-corrected prevalence of 'incorrect'
      with a bootstrap CI.

Design: among the 1,171 definitively-coded papers, the classifier called
  - N_neg = 37 papers 'not-incorrect' (CORRECT/ACCEPTABLE) -> ALL 37 human-coded
            (a census of this stratum -> contributes NO sampling uncertainty);
  - N_pos = 1,134 papers 'incorrect' (INCORRECT/MIXED) -> 63 human-coded
            (a sample -> the only source of sampling uncertainty).

Run: python3 validation_correct.py [path_to_filled_sheet.xlsx|csv]
"""
import sys, csv, json, random, statistics

random.seed(2025)
N_DEF = 1171  # definitively-coded papers (the census)

def human_binary(code):
    code = (code or "").strip().upper()
    if code in ("INCORRECT", "MIXED"):
        return "incorrect"
    if code in ("CORRECT", "ACCEPTABLE"):
        return "ok"
    return "unclear"   # UNCLEAR / NOT_RELEVANT / blank

def clf_binary(code):
    return "incorrect" if code in ("INCORRECT", "MIXED") else "ok"

# ---- load filled sheet (xlsx or csv) -> {id: human_code} ----
path = sys.argv[1] if len(sys.argv) > 1 else "validation_coding_sheet.xlsx"
human = {}
if path.endswith(".xlsx"):
    from openpyxl import load_workbook
    ws = load_workbook(path)["coding"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(x) for x in rows[0]]
    id_i, code_i = hdr.index("id"), hdr.index("CODE")
    for r in rows[1:]:
        if r[id_i] is not None:
            human[str(r[id_i])] = (r[code_i] or "")
else:
    for r in csv.DictReader(open(path)):
        human[r["id"]] = r["CODE"]

design = json.load(open("validation_sample_design.json"))
AUTO = {r["id"]: r["auto_code"] for r in csv.DictReader(open("coded_auto.csv"))}

coded = {i: c for i, c in human.items() if str(c).strip()}
print(f"Coded rows: {len(coded)}/{len(human)}")
from collections import Counter
print("Human code distribution:", dict(Counter(str(c).strip().upper() for c in coded.values())))
n_unclear = sum(1 for c in coded.values() if human_binary(c) == "unclear")
print(f"Human UNCLEAR/NOT_RELEVANT: {n_unclear}")

# ---- split by classifier stratum ----
neg = [(i, human_binary(c)) for i, c in coded.items() if clf_binary(AUTO[i]) == "ok"]
pos = [(i, human_binary(c)) for i, c in coded.items() if clf_binary(AUTO[i]) == "incorrect"]
print(f"\nClassifier-negative stratum coded: {len(neg)} (of {design['n_neg_total']} in corpus)")
print(f"Classifier-positive stratum coded: {len(pos)} (of {design['n_pos_total']} in corpus)")

# ---- (1) corpus-weighted agreement + kappa (binary) ----
Nneg, Npos = design["n_neg_total"], design["n_pos_total"]
def rate(stratum, want):  # fraction of stratum with human==want (definitive only)
    d = [h for _, h in stratum if h != "unclear"]
    return (sum(1 for h in d if h == want) / len(d)) if d else float("nan")
# weighted confusion proportions over the corpus
w_neg, w_pos = Nneg / N_DEF, Npos / N_DEF
# P(clf, human) cells
p_neg_ok  = w_neg * rate(neg, "ok");        p_neg_inc = w_neg * rate(neg, "incorrect")
p_pos_inc = w_pos * rate(pos, "incorrect"); p_pos_ok  = w_pos * rate(pos, "ok")
po = p_neg_ok + p_pos_inc                    # agreement (clf ok & human ok) + (clf inc & human inc)
# marginals
clf_inc = w_pos; clf_ok = w_neg
hum_inc = p_neg_inc + p_pos_inc; hum_ok = p_neg_ok + p_pos_ok
pe = clf_inc * hum_inc + clf_ok * hum_ok
kappa = (po - pe) / (1 - pe) if pe < 1 else float("nan")
print(f"\n(1) Corpus-weighted binary agreement: {100*po:.1f}%   Cohen's kappa: {kappa:.3f}")

# ---- (2) design-based corrected prevalence ----
# a = true-incorrect count in the fully-observed negative stratum (census, fixed)
a = sum(1 for _, h in neg if h == "incorrect")
pos_def = [h for _, h in pos if h != "unclear"]
b_rate = sum(1 for h in pos_def if h == "incorrect") / len(pos_def)   # incorrect rate among positives
T = a + Npos * b_rate
p_corr = T / N_DEF
# bootstrap: negatives fixed (census); resample the positive sample only
boot = []
for _ in range(20000):
    bs = [random.choice(pos_def) for _ in pos_def]
    br = sum(1 for h in bs if h == "incorrect") / len(bs)
    boot.append((a + Npos * br) / N_DEF)
boot.sort()
lo, hi = boot[int(0.025*len(boot))], boot[int(0.975*len(boot))]
naive = 1134 / N_DEF
print(f"\n(2) Naive classifier prevalence:        {100*naive:.1f}%  (1134/{N_DEF})")
print(f"    False-negatives in neg stratum (a): {a}/{len(neg)} truly incorrect")
print(f"    Incorrect rate among positives:     {100*b_rate:.1f}%  ({sum(1 for h in pos_def if h=='incorrect')}/{len(pos_def)})")
print(f"    Misclassification-corrected prevalence: {100*p_corr:.1f}%")
print(f"    95% bootstrap CI (positive-stratum):   [{100*lo:.1f}, {100*hi:.1f}]")
print("\nNote: the negative stratum is a census (all classifier-negatives, human-coded),")
print("so it adds no sampling uncertainty; the CI reflects only the positive-stratum sample.")
