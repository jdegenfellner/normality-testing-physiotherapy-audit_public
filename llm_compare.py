#!/usr/bin/env python3
"""Compare the LLM classifier against (a) the rule-based classifier over the whole
corpus and (b) the human gold standard on the 100-paper validation sample, used
as an independent second coder. Agreement with the human is corpus-weighted on the
SAME stratification as validation_analyze.py, so the rule-based row reproduces the
90.7% / AC1 0.89 reported there; the LLM row is directly comparable."""
import csv
from collections import Counter
from openpyxl import load_workbook

def b(c):
    c = str(c).strip().upper()
    return "INC" if c in ("INCORRECT", "MIXED") else "OK" if c in ("CORRECT", "ACCEPTABLE") else None

# ---- load the three code sources ----
corpus = [r for r in csv.DictReader(open("coded_auto.csv"))
          if r["year"].isdigit() and 2015 <= int(r["year"]) <= 2024]
rule = {r["id"]: r["auto_code"] for r in corpus if r["auto_code"] != "UNCLEAR"}
N_DEF = sum(1 for r in corpus if r["auto_code"] in ("CORRECT", "ACCEPTABLE", "INCORRECT", "MIXED"))
N_pos = sum(1 for r in corpus if r["auto_code"] in ("INCORRECT", "MIXED"))
N_neg = N_DEF - N_pos
llm = {r["id"]: r["llm_code"] for r in csv.DictReader(open("llm_codes.csv"))}
ws = load_workbook("validation_coding_sheet.xlsx").active
hdr = [c.value for c in ws[1]]; i_id = hdr.index("id"); i_c = hdr.index("CODE")
human = {str(r[i_id].value): r[i_c].value for r in ws.iter_rows(min_row=2) if r[i_id].value}

defin = [i for i in rule if i in llm]

# ---- (A) LLM vs rule-based over the whole corpus ----
print(f"definitive papers (rule & LLM): {len(defin)}")
raw = sum(rule[i] == llm[i] for i in defin) / len(defin)
li = sum(b(llm[i]) == "INC" for i in defin)
ld = sum(b(llm[i]) in ("INC", "OK") for i in defin)
def ac1_unweighted(pairs):  # for the full census (no stratification)
    pairs = [(x, y) for x, y in pairs if x and y]
    n = len(pairs); po = sum(x == y for x, y in pairs) / n
    pi_inc = sum(x == "INC" for p in pairs for x in p) / (2 * n)
    pe = 2 * pi_inc * (1 - pi_inc)
    return 100 * po, (po - pe) / (1 - pe), n

print("\n(A) LLM vs RULE-BASED, full corpus")
print(f"    raw 5-class agreement: {100*raw:.1f}%")
po, a1, n = ac1_unweighted([(b(rule[i]), b(llm[i])) for i in defin])
print(f"    binary (incorrect vs not) agreement: {po:.1f}%  AC1 {a1:.3f}  (n={n} determinable)")
print(f"    LLM standalone prevalence: {100*li/len(defin):.1f}% of definitive | "
      f"{100*li/ld:.1f}% of determinable  (rule-based naive 96.8%)")
print("    LLM code distribution:", dict(Counter(llm[i] for i in defin)))

# ---- corpus-weighted agreement with the human, same strata as the correction ----
def weighted(rater):
    w_neg, w_pos = N_neg / N_DEF, N_pos / N_DEF
    strata = {"neg": [i for i in human if i in rule and b(rule[i]) == "OK"],
              "pos": [i for i in human if i in rule and b(rule[i]) == "INC"]}
    p = Counter()
    for name, w in (("neg", w_neg), ("pos", w_pos)):
        ids = [i for i in strata[name] if b(human[i]) and b(rater.get(i))]
        n = len(ids)
        for i in ids:
            p[(b(human[i]), b(rater[i]))] += w / n
    po = p[("OK", "OK")] + p[("INC", "INC")]
    hum_inc = p[("INC", "INC")] + p[("INC", "OK")]
    rat_inc = p[("INC", "INC")] + p[("OK", "INC")]
    pi_inc = (hum_inc + rat_inc) / 2; pi_ok = 1 - pi_inc
    pe = 2 * pi_ok * pi_inc
    return 100 * po, (po - pe) / (1 - pe)

print("\n(B) Agreement with the human gold standard (corpus-weighted, 100-paper sample)")
print("    human vs RULE-BASED: %.1f%%  AC1 %.3f   (reproduces validation_analyze.py)" % weighted(rule))
print("    human vs LLM:        %.1f%%  AC1 %.3f" % weighted(llm))

# where LLM and rule disagree on the binary call, who matches the human?
dis = [i for i in human if i in llm and b(rule.get(i)) and b(llm[i]) and b(rule[i]) != b(llm[i])
       and b(human[i])]
right = sum(b(llm[i]) == b(human[i]) for i in dis)
print(f"\n    where LLM & rule-based disagree (binary, n={len(dis)}): LLM matches human in {right}")

# 5-class confusion, human (rows) x LLM (cols)
cats = ["CORRECT", "ACCEPTABLE", "INCORRECT", "MIXED", "UNCLEAR"]
print("\n    Confusion human (rows) x LLM (cols):")
print("    " + "".join(f"{c[:5]:>7}" for c in cats))
for hc in cats:
    row = [sum(1 for i in human if i in llm and str(human[i]).upper() == hc
               and str(llm[i]).upper() == lc) for lc in cats]
    if sum(row):
        print(f"    {hc[:9]:9s}" + "".join(f"{x:7d}" for x in row))
