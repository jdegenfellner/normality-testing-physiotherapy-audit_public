#!/usr/bin/env python3
"""Convert validation_coding_sheet.csv to a fill-in-friendly .xlsx:
a CODE dropdown (data validation), frozen header, wrapped text, plus an
instructions sheet."""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

CODES = ["CORRECT", "ACCEPTABLE", "INCORRECT", "MIXED", "UNCLEAR", "NOT_RELEVANT"]
rows = list(csv.reader(open("validation_coding_sheet.csv")))
header, data = rows[0], rows[1:]

wb = Workbook()
ws = wb.active
ws.title = "coding"

hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(bold=True, color="FFFFFF")
code_fill = PatternFill("solid", fgColor="FFF2CC")
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(header)
for c in range(1, len(header) + 1):
    cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(vertical="center")

for r in data:
    ws.append(r)

widths = {"idx": 5, "id": 12, "journal": 16, "year": 7, "title": 42,
          "normality_sentences": 95, "CODE": 16}
for i, name in enumerate(header, 1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)

code_col = header.index("CODE") + 1
code_letter = get_column_letter(code_col)
wrap_cols = [header.index("title") + 1, header.index("normality_sentences") + 1]
for r in range(2, len(data) + 2):
    for c in wrap_cols:
        ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    cc = ws.cell(r, code_col); cc.fill = code_fill
    cc.alignment = Alignment(horizontal="center", vertical="center")
    cc.border = border

dv = DataValidation(type="list", formula1='"%s"' % ",".join(CODES),
                    allow_blank=True, showDropDown=False)
dv.error = "Please pick one of: " + ", ".join(CODES)
dv.prompt = "Pick the code that matches what the normality test was applied to."
ws.add_data_validation(dv)
dv.add(f"{code_letter}2:{code_letter}{len(data)+1}")

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 20

# --- instructions sheet ---
wsi = wb.create_sheet("instructions")
instr = [
    ("What was the normality test applied to? Put ONE code per row in the 'CODE' column of the 'coding' tab.", True),
    ("", False),
    ("CORRECT      - applied to the model RESIDUALS.", False),
    ("ACCEPTABLE   - applied SEPARATELY WITHIN EACH GROUP (= residuals for a balanced 2-group design).", False),
    ("INCORRECT    - applied to the RAW DATA / outcome / variables / distribution, with no residual or", False),
    ("               per-group qualifier. An unqualified 'data were tested for normality' is INCORRECT.", False),
    ("MIXED        - both a correct/per-group AND an unqualified pooled statement in the same paper.", False),
    ("UNCLEAR      - a test is mentioned but the target cannot be determined / the sentence is not a normality test.", False),
    ("NOT_RELEVANT - 'normal'/'distribution' wording unrelated to a model assumption (Gaussian filter, normal range...).", False),
    ("", False),
    ("Borderline rules:", True),
    ("- 'data were Shapiro-Wilk tested and found normally distributed'  -> INCORRECT", False),
    ("- 'each group was assessed for normality'  -> ACCEPTABLE", False),
    ("- group words that refer to the COMPARISON test (not to normality testing)  -> still INCORRECT", False),
    ("- ambiguous / only vague wording  -> UNCLEAR", False),
    ("", False),
    ("Code blind: do not look anything up unless a sentence is genuinely ambiguous. The relevant", False),
    ("normality sentence(s) are already provided in the 'normality_sentences' column.", False),
]
for i, (text, bold) in enumerate(instr, 1):
    cell = wsi.cell(i, 1, text)
    if bold:
        cell.font = Font(bold=True)
wsi.column_dimensions["A"].width = 115

wb.save("validation_coding_sheet.xlsx")
print(f"Wrote validation_coding_sheet.xlsx ({len(data)} rows, CODE dropdown, instructions tab).")
