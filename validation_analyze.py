#!/usr/bin/env python3
"""Full validation analysis: confusion matrix + design-based misclassification
correction with bootstrap CIs. Stratified sample: all N_neg classifier-negatives
(census) + a random sample of classifier-positives."""
import csv, json, random
from collections import Counter
from openpyxl import load_workbook
random.seed(2025)

# Restrict to full publication years 2015-2024; online-first 2025 records that
# the search returned are excluded to match the stated study window. (All 2025
# records are classifier-incorrect and none fall in the validation sample, so the
# confusion matrix is unaffected; only the positive-stratum totals shrink.)
_corpus = [r for r in csv.DictReader(open("coded_auto.csv"))
           if 2015 <= int(r["year"]) <= 2024]
AUTO = {r["id"]: r["auto_code"] for r in _corpus}
_isdef = lambda c: c in ("CORRECT", "ACCEPTABLE", "INCORRECT", "MIXED")
_isinc = lambda c: c in ("INCORRECT", "MIXED")
N_DEF = sum(_isdef(r["auto_code"]) for r in _corpus)
N_pos = sum(_isinc(r["auto_code"]) for r in _corpus)
N_neg = N_DEF - N_pos

ws = load_workbook("validation_coding_sheet.xlsx")["coding"]
hdr = [c.value for c in ws[1]]
ii, ci = hdr.index("id"), hdr.index("CODE")
human = {str(ws.cell(r, ii + 1).value): str(ws.cell(r, ci + 1).value).strip().upper()
         for r in range(2, ws.max_row + 1) if ws.cell(r, ii + 1).value}

def hb(c):  # human binary
    return "INC" if c in ("INCORRECT", "MIXED") else ("OK" if c in ("CORRECT", "ACCEPTABLE") else "UNC")
def cb(c):
    return "INC" if c in ("INCORRECT", "MIXED") else "OK"

# ---- confusion matrix (raw, stratified sample) ----
# Human and classifier share the same five levels; any "not relevant" judgement
# (matched text that is not a genuine normality test) is folded into UNCLEAR so
# the two schemes are identical.
hud = lambda c: "UNCLEAR" if c in ("UNCLEAR", "NOT_RELEVANT") else c
cats = ["CORRECT", "ACCEPTABLE", "INCORRECT", "MIXED", "UNCLEAR"]
clf_cats = ["CORRECT", "ACCEPTABLE", "INCORRECT", "MIXED"]
conf = Counter((AUTO[i], hud(human[i])) for i in human)
print("CONFUSION (rows = classifier, cols = human); stratified sample")
print(f"{'clf \\\\ human':14}" + "".join(f"{c[:5]:>7}" for c in cats) + "   tot")
for cc in clf_cats:
    row = [conf[(cc, hc)] for hc in cats]
    print(f"{cc:14}" + "".join(f"{x:>7}" for x in row) + f"   {sum(row)}")

# ---- per-stratum human breakdown ----
neg = [human[i] for i in human if cb(AUTO[i]) == "OK"]
pos = [human[i] for i in human if cb(AUTO[i]) == "INC"]
print(f"\nNEG stratum (classifier CORRECT/ACCEPTABLE), all {len(neg)} (census):", dict(Counter(hb(c) for c in neg)))
print(f"POS stratum (classifier INCORRECT/MIXED), {len(pos)} of {N_pos} sampled:", dict(Counter(hb(c) for c in pos)))

# ---- corpus-weighted binary agreement + chance-corrected coefficients ----
# Primary measure is Gwet's AC1 (robust to the extreme class imbalance, which
# deflates Cohen's kappa via the prevalence paradox). PABAK and kappa printed too.
def weighted_agreement():
    w_neg, w_pos = N_neg / N_DEF, N_pos / N_DEF
    # cells P(clf,human-binary), human UNC excluded -> renormalise within determinable
    def br(strat):  # binary rates among determinable
        d = [hb(c) for c in strat if hb(c) != "UNC"]
        return Counter(d), len(d)
    cn, nn = br(neg); cp, npd = br(pos)
    p = {}
    p[("OK","OK")]   = w_neg*cn["OK"]/nn
    p[("OK","INC")]  = w_neg*cn["INC"]/nn
    p[("INC","INC")] = w_pos*cp["INC"]/npd
    p[("INC","OK")]  = w_pos*cp["OK"]/npd
    po = p[("OK","OK")] + p[("INC","INC")]
    clf_ok, clf_inc = w_neg, w_pos
    hum_ok = p[("OK","OK")]+p[("INC","OK")]; hum_inc = p[("OK","INC")]+p[("INC","INC")]
    # Cohen's kappa
    pe_k = clf_ok*hum_ok + clf_inc*hum_inc
    kappa = (po-pe_k)/(1-pe_k)
    # PABAK = 2*po - 1 (chance term fixed at 0.5)
    pabak = 2*po - 1
    # Gwet's AC1: pe = (1/(q-1)) * sum_k pi_k(1-pi_k); for q=2 -> 2*pi_ok*pi_inc
    pi_ok  = (clf_ok + hum_ok)/2
    pi_inc = (clf_inc + hum_inc)/2
    pe_g = 2*pi_ok*pi_inc
    ac1 = (po-pe_g)/(1-pe_g)
    return po, ac1, pabak, kappa
po, ac1, pabak, kap = weighted_agreement()
print(f"\nCorpus-weighted binary agreement (determinable): {100*po:.1f}%")
print(f"  Gwet's AC1 (primary): {ac1:.3f}   PABAK: {pabak:.3f}   Cohen's kappa: {kap:.3f}")

# ---- design-based corpus category totals + two prevalence estimands ----
def estimate(neg_codes, pos_codes):
    cn = Counter(hb(c) for c in neg_codes)             # census (exact)
    cp = Counter(hb(c) for c in pos_codes); n = len(pos_codes)
    scale = N_pos / n
    inc = cn["INC"] + scale*cp["INC"]
    ok  = cn["OK"]  + scale*cp["OK"]
    unc = cn["UNC"] + scale*cp["UNC"]
    pA = inc/(inc+ok+unc)            # incorrect / all definitive
    pB = inc/(inc+ok)               # incorrect / determinable (excl. unclear/not-relevant)
    return inc, ok, unc, pA, pB
inc, ok, unc, pA, pB = estimate(neg, pos)
# bootstrap (negatives census-fixed; resample positives)
bA, bB = [], []
for _ in range(20000):
    bs = [random.choice(pos) for _ in pos]
    i,o,u,a,b = estimate(neg, bs)
    bA.append(a); bB.append(b)
bA.sort(); bB.sort()
# 95% bootstrap CI: central 95%, i.e. 2.5% in each tail (percentiles 0.025 / 0.975)
ci = lambda v: (v[int(.025*len(v))], v[int(.975*len(v))])
loA,hiA = ci(bA); loB,hiB = ci(bB)
print(f"\nDesign-based corpus estimates (definitive={N_DEF}):")
print(f"  est. incorrect={inc:.0f}, correct/acceptable={ok:.0f}, unclear/not-relevant={unc:.0f}")
print(f"  Estimand A  (incorrect / all definitive):     {100*pA:.1f}%   95% CI [{100*loA:.1f}, {100*hiA:.1f}]")
print(f"  Estimand B  (incorrect / determinable):       {100*pB:.1f}%   95% CI [{100*loB:.1f}, {100*hiB:.1f}]")
print(f"  (naive classifier prevalence was 96.8%)")

# ---- export the numbers used by the estimand-explainer figure (estimand_fig.R) ----
import json as _json
_total = N_DEF
_inc   = round(inc); _ind = round(unc); _ok = _total - _inc - _ind   # ok set so segments sum to total
_fig = {
    "total": _total, "incorrect": _inc, "ok": _ok, "indeterminate": _ind,
    "determinable": _total - _ind,
    "pA": round(100*pA, 1), "loA": round(100*loA, 1), "hiA": round(100*hiA, 1),
    "pB": round(100*pB, 1), "loB": round(100*loB, 1), "hiB": round(100*hiB, 1),
    "raw_inc": N_pos, "raw_pct": round(100*N_pos/N_DEF, 1),
}
_json.dump(_fig, open("estimand_figure_data.json", "w"), indent=2)
print("wrote estimand_figure_data.json")
